from __future__ import annotations

import argparse
from datetime import datetime, time, timedelta
import json
from pathlib import Path

from openpyxl import load_workbook

from analysis_lib.player_card import load_confirmed_players


def seconds(value: object) -> float:
    if isinstance(value, timedelta):
        return value.total_seconds()
    if isinstance(value, datetime):
        value = value.time()
    if isinstance(value, time):
        return value.hour * 3600 + value.minute * 60 + value.second + value.microsecond / 1_000_000
    if isinstance(value, (int, float)):
        # Excel times are stored as a fraction of one day.
        return float(value) * 86400 if 0 <= float(value) < 1 else float(value)
    parts = str(value).strip().split(":")
    if len(parts) == 3:
        return int(parts[0]) * 3600 + int(parts[1]) * 60 + float(parts[2])
    if len(parts) == 2:
        return int(parts[0]) * 60 + float(parts[1])
    raise ValueError(f"无法解析时间: {value!r}")


def event_type(text: str) -> str:
    if "射门" in text:
        return "shot"
    if "解围" in text:
        return "clearance"
    if any(key in text for key in ("传球", "传递", "送出", "转移")):
        return "pass_or_distribution"
    if any(key in text for key in ("抢断", "丢球", "失去球权", "被断", "拦截")):
        return "possession_loss_or_tackle"
    if any(key in text for key in ("盘带", "推进", "带球")):
        return "dribble"
    if "护球" in text:
        return "ball_protection"
    return "key_action"


def read_ledger(path: Path) -> list[dict]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["标注台账"]
    rows = sheet.iter_rows(values_only=True)
    next(rows)  # merged title
    headers = [str(value).strip() if value is not None else "" for value in next(rows)]
    result = []
    for values in rows:
        row = dict(zip(headers, values))
        if row.get("切片ID"):
            result.append(row)
    workbook.close()
    return result


def convert(*, xlsx: Path, detector_events: Path, clips_dir: Path,
            numbers: Path, output: Path, max_match_seconds: float) -> dict:
    if output.exists():
        raise FileExistsError(f"输出文件已存在，拒绝覆盖: {output}")
    clips_dir = clips_dir.resolve()
    ledger = read_ledger(xlsx)
    detector_data = json.loads(detector_events.read_text(encoding="utf-8"))
    detector = detector_data.get("events", []) if isinstance(detector_data, dict) else detector_data
    players, audit = load_confirmed_players(numbers)
    gid_identity = {gid: player for player in players.values() for gid in player["global_ids"]}
    result = []
    for index, row in enumerate(ledger, 1):
        anchor = seconds(row["锚点时间"])
        nearest = min(detector, key=lambda event: abs(float(event["event_time_seconds"]) - anchor)) if detector else None
        delta = abs(float(nearest["event_time_seconds"]) - anchor) if nearest else None
        matched = nearest if delta is not None and delta <= max_match_seconds else None
        candidates = matched.get("actor_candidates", []) if matched else []
        primary = matched.get("primary_global_id") if matched else None
        secondary = next((candidate.get("global_id") for candidate in candidates
                          if candidate.get("global_id") != primary), None)
        identity = gid_identity.get(int(primary)) if primary is not None else None
        event_id = str(row["切片ID"])
        clip_name = str(row.get("切片文件") or f"{event_id}.mp4")
        clip = (clips_dir / clip_name).resolve()
        if not clip.is_file():
            raise FileNotFoundError(f"缺少标注切片: {clip}")
        description = " | ".join(str(row.get(key) or "").strip() for key in
                                 ("事件类型提示", "行为标签（1～3个）", "备注") if row.get(key))
        result.append({
            "event_id": event_id,
            "start_time": round(seconds(row["起始时间戳"]), 3),
            "end_time": round(seconds(row["结束时间戳"]), 3),
            "primary_global_id": int(primary) if primary is not None else None,
            "secondary_global_id": int(secondary) if secondary is not None else None,
            "event_type": event_type(description),
            "video_anchor_path": str(clip),
            "jersey_number": identity["jersey_number"] if identity else None,
            "confidence": round(float(matched.get("score", 0)) / 100, 4) if matched else 0.0,
            "anchor_time": round(anchor, 3),
            "description": description,
            "main_dimension_seed": row.get("主维度"),
            "uncertainty": row.get("拿不准"),
            "actor_assignment_status": "provisional_nearest_detector_event" if matched else "unassigned",
            "actor_match_delta_sec": round(delta, 3) if delta is not None else None,
            "identity_status": "eligible_confirmed" if identity else "not_confirmed",
            "player_id": identity["player_id"] if identity else None,
        })
    payload = {
        "schema_version": "events-for-annotation-v1",
        "total_events": len(result),
        "events": result,
        "provenance": {
            "xlsx": str(xlsx.resolve()),
            "detector_events": str(detector_events.resolve()),
            "identity_verifier": str(numbers.resolve()),
            "actor_policy": f"nearest detector event within {max_match_seconds:g}s; provisional only",
            "identity_policy": audit["policy"],
        },
    }
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return payload


def main() -> None:
    parser = argparse.ArgumentParser(description="把事件标注台账转换为标准 events_for_annotation.json")
    parser.add_argument("--xlsx", type=Path, required=True)
    parser.add_argument("--detector-events", type=Path, required=True)
    parser.add_argument("--clips-dir", type=Path, required=True)
    parser.add_argument("--numbers", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--max-match-seconds", type=float, default=8.0)
    args = parser.parse_args()
    payload = convert(xlsx=args.xlsx, detector_events=args.detector_events,
                      clips_dir=args.clips_dir, numbers=args.numbers,
                      output=args.output, max_match_seconds=args.max_match_seconds)
    print(json.dumps({"output": str(args.output.resolve()), "events": payload["total_events"]}, ensure_ascii=False))


if __name__ == "__main__":
    main()
